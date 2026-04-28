#!/usr/bin/env python3
"""Read SDD v1.1.0 and extract DbDaemon API specifications."""

import openpyxl
import sys

# Read the SDD v1.1.0
try:
    wb = openpyxl.load_workbook('docs/FSS_SoftwareDetailedDesign_v1.1.0.xlsx')
    print("=" * 80)
    print("SHEET NAMES IN SDD v1.1.0:")
    print("=" * 80)
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"{i}. {sheet}")
    
    # Look for DbDaemon content
    print("\n" + "=" * 80)
    print("SEARCHING FOR DB_DAEMON CONTENT:")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        content = []
        for row in ws.iter_rows(values_only=True):
            content.append(row)
        
        # Convert to string and check for DbDaemon references
        sheet_str = str(content).lower()
        if 'db' in sheet_str or 'daemon' in sheet_str or 'database' in sheet_str:
            print(f"\n--- Sheet: {sheet_name} ---")
            for i, row in enumerate(content[:50], 1):
                if any(row):
                    print(f"Row {i}: {row}")
            
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
