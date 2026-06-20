#!/usr/bin/env python3
"""Comprehensive fixes for all 3 FSS docs:
1. Apply Google Sans font to ALL text
2. Delete strikethrough rows in SYS
3. Consolidate SYS ChangeLog v1.2.0 entries
4. Update SWE old module names and content
"""

import os, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter

DOCS = os.path.join(os.path.dirname(__file__), '..', 'docs')

# Google Sans font definition
gst = Font(name='Google Sans Text', size=11, color=Color(theme=1))
gst_bold = Font(name='Google Sans Text', size=11, bold=True, color=Color(theme=1))
gst_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
gst_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_fill = PatternFill(fill_type='solid', fgColor=Color(theme=5, tint=0.7999816888943144))


def apply_gst_to_all_cells(ws):
    """Apply Google Sans Text font to every cell in the worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                was_bold = cell.font and cell.font.bold
                cell.font = gst_bold if was_bold else gst


def delete_row(ws, row):
    """Delete a row by shifting cells up (preserving merge)."""
    max_row = ws.max_row
    max_col = ws.max_column

    # Move cells up
    for r in range(row, max_row):
        for c in range(1, max_col + 1):
            src = ws.cell(row=r + 1, column=c)
            dst = ws.cell(row=r, column=c)
            dst.value = src.value
            dst.font = copy.copy(src.font)
            dst.alignment = copy.copy(src.alignment)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)

    # Clear last row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=max_row, column=c)
        cell.value = None
        cell.font = gst
        cell.alignment = None
        cell.fill = PatternFill()
        cell.border = Border()


# ============================================================
# 1. SDD v1.2.0
# ============================================================
path = os.path.join(DOCS, 'FSS_SoftwareDetailedDesign_v1.2.0.xlsx')
wb = load_workbook(path)

# Apply Google Sans to all cells in every sheet
for sn in wb.sheetnames:
    ws = wb[sn]
    apply_gst_to_all_cells(ws)
    # Also set column widths to accommodate text
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if letter in ws.column_dimensions:
            current = ws.column_dimensions[letter].width or 15
            if current < 15:
                ws.column_dimensions[letter].width = 20

wb.save(path)
print(f'SDD v1.2.0: Google Sans applied')

# ============================================================
# 2. SWE v1.3.0
# ============================================================
path = os.path.join(DOCS, 'FSS_SoftwareEngineering_v1.3.0.xlsx')
wb = load_workbook(path)

ws = wb['1. SoftwareReqSpec']

# Update heading names and allocations (old -> new)
updates = {
    # Row 7: MMM-Env-Supervisor heading
    7: {
        5: 'MMM-FSS-Env (Environment Monitoring) - Quản lý đọc cảm biến môi trường (SHT3x, VL53L0X) qua C++ SensorDaemon',
    },
    # Row 8: SW.ENV.01 - allocation
    8: {9: 'sensor_daemon/src/SensorDaemonMain.cpp (I2cHandler)'},
    # Row 9: SW.ENV.02
    9: {9: 'sensor_daemon/src/InputProcessor.cpp'},
    # Row 10: SW.ENV.03
    10: {9: 'sensor_daemon/src/Sht3xDriver.cpp'},
    # Row 11: SW.ENV.04
    11: {
        5: 'Presence Detection',
        9: 'sensor_daemon/src/Vl53l0xDriver.cpp',
    },
    # Row 12: SW.ENV.05 - Power Management (vcgencmd is now handled by electron)
    12: {9: 'electron_app/magicmirror/modules/MMM-FSS-Env/node_helper.js'},
    # Row 13: SW.ENV.06
    13: {9: 'sensor_daemon/src/OutputProcessor.cpp (D-Bus)'},
    # Row 14: SW.ENV.07
    14: {9: 'db_daemon/src/SqliteManager.py'},
    # Row 15: MMM-Food-Reco heading
    15: {
        5: 'FRTApp (Food Recognition & Tracking) - Pipeline AI thời gian thực (C++ camera + Python AI)',
    },
    # Row 16: SW.AI.01
    16: {9: 'frt_app/py_ai_core/src/main.py (FrtDbusInterface)'},
    # Row 17: SW.AI.02
    17: {9: 'frt_app/py_ai_core/src/YoloPipeline.py'},
    # Row 18: SW.AI.03
    18: {9: 'frt_app/cpp_camera_core/src/main.cpp (V4L2)'},
    # Row 19: SW.AI.04 - update description (Ultralytics -> tflite-runtime)
    19: {
        6: 'The software shall load the YOLOv11 TFLite model and ByteTrack tracker to perform real-time multi-object detection and tracking on video frames.',
        9: 'frt_app/py_ai_core/src/YoloTfliteEngine.py + ByteTracker.py',
    },
    # Row 20: SW.AI.05 - update description
    20: {
        6: 'The software shall filter detection results to include only defined food classes (custom label mapping).',
        9: 'frt_app/py_ai_core/src/YoloPipeline.py',
    },
    # Row 21: SW.AI.06
    21: {9: 'frt_app/py_ai_core/src/YoloTfliteEngine.py'},
    # Row 22: SW.AI.07 - update description (direction IN/OUT)
    22: {
        6: 'The software shall generate a JSON object containing tracking events with direction (IN/OUT) using VirtualLineDetector and ByteTrack.',
        9: 'frt_app/py_ai_core/src/VirtualLineDetector.py',
    },
    # Row 23: SW.AI.08 - update description
    23: {
        6: 'The software shall save annotated frames (with bounding boxes) to /opt/fss/latest_preview.jpg for UI consumption.',
        9: 'frt_app/py_ai_core/src/main.py',
    },
    # Row 24: MMM-Database heading
    24: {
        5: 'DBDaemon (Data Controller) - Quản lý 3 SQLite databases (fss_data.db, FSS_Inventory.db, FSS_Request.db)',
    },
    # Row 25: SW.DB.01
    25: {9: 'db_daemon/src/SqliteManager.py'},
    # Row 26: SW.DB.02
    26: {9: 'db_daemon/src/DbDaemonMain.py'},
    # Row 27: SW.DB.03
    27: {
        6: 'The software shall execute a transactional update for inventory changes upon DOOR_CLOSE event via D-Bus FoodDetected signal from FRTApp.',
        9: 'db_daemon/src/DbDaemonMain.py (process_food_tracking_event)',
    },
    # Row 28: SW.DB.04
    28: {9: 'db_daemon/src/DbDbusInterface.py (GetInventory method)'},
    # Row 29: SW.DB.05
    29: {9: 'db_daemon/src/SqliteManager.py (WAL mode)'},
    # Row 30: MMM-Food-Management heading
    30: {
        5: 'MMM-FSS-Inventory (User Interface) - Hiển thị danh sách thực phẩm trên MagicMirror',
    },
    # Row 31: SW.UI.01
    31: {9: 'electron_app/magicmirror/modules/MMM-FSS-Inventory/MMM-FSS-Inventory.js'},
    # Row 32: SW.UI.02
    32: {9: 'electron_app/magicmirror/modules/MMM-FSS-Inventory/MMM-FSS-Inventory.js'},
    # Row 33: SW.UI.03
    33: {9: 'electron_app/magicmirror/modules/MMM-FSS-Inventory/MMM-FSS-Inventory.js'},
    # Row 34: SW.UI.04
    34: {9: 'electron_app/magicmirror/modules/MMM-FSS-LivePreview/MMM-FSS-LivePreview.js'},
    # Row 35: SW.UI.05
    35: {9: 'electron_app/magicmirror/modules/MMM-FSS-Inventory/MMM-FSS-Inventory.css'},
}

for row, cols in updates.items():
    for col, val in cols.items():
        ws.cell(row=row, column=col).value = val

# Apply Google Sans to all sheets
for sn in wb.sheetnames:
    apply_gst_to_all_cells(wb[sn])

wb.save(path)
print(f'SWE v1.3.0: Content updated + Google Sans applied')

# ============================================================
# 3. SYS v1.2.0
# ============================================================
path = os.path.join(DOCS, 'FSS_SystemEngineering_v1.2.0.xlsx')
wb = load_workbook(path)

# ---- Fix A: Consolidate ChangeLog duplicate v1.2.0 entries ----
ws_cl = wb['ChangeLog']
v120_rows = []
for row in range(1, ws_cl.max_row + 1):
    if ws_cl.cell(row=row, column=1).value == '1.2.0':
        v120_rows.append(row)

if len(v120_rows) >= 2:
    # Original entry at v120_rows[0] (old planned changes)
    # New entry at v120_rows[1] (my actual update)
    # Merge them
    old_desc = str(ws_cl.cell(row=v120_rows[0], column=2).value or '')
    new_desc = str(ws_cl.cell(row=v120_rows[1], column=2).value or '')

    merged = (
        '[Update]: Removed LED Module and Relay dependencies due to hardware scope change\n'
        '[Update]: Updated Hardware Elements. Changed Door Sensor to MC-38 (GPIO), Presence Sensor to VL53L0X (I2C)\n'
        '[Update]: Aligned System Requirements with FRT Pipeline v1.2 (continuous tracking during DOOR_OPEN)\n'
        '[Update]: Changed camera hardware from Pi Camera (CSI) to USB Camera (UVC)\n'
        '[Update]: [1.ReqSpec] [All] Added SYS.FUNC.AI.02 for recommendation and extraction\n'
        '[Update]: [2.Elements] [All] Added SYS.ELEM.11-13 for RecommendDaemon, RecipeExtractor, C TFLite Reader\n'
        '[Update]: [3.Interfaces] [All] Added SYS.IF.INT.03-06 for new IPC flows\n'
        '[Update]: [4.Modes] [All] Removed Relay/LED actuation from operation modes\n'
        '[Fix]: Version consistency - ReqSpec sheet v0.6.0 updated to v1.2.0'
    )

    # Keep the first entry (row v120_rows[0]), update its description
    ws_cl.cell(row=v120_rows[0], column=2).value = merged

    # Delete the second entry row (v120_rows[1])
    delete_row(ws_cl, v120_rows[1])

# ---- Fix B: Delete strikethrough rows in 1. ReqSpec ----
ws_req = wb['1. ReqSpec']
# Find and delete rows with strikethrough
rows_to_delete = []
for row in range(1, ws_req.max_row + 1):
    for col in range(1, ws_req.max_column + 1):
        cell = ws_req.cell(row=row, column=col)
        if cell.font and cell.font.strike:
            rows_to_delete.append(row)
            break

# Delete from bottom to top to preserve row numbers
for row in sorted(rows_to_delete, reverse=True):
    delete_row(ws_req, row)

# ---- Fix C: Apply Google Sans to all sheets ----
for sn in wb.sheetnames:
    apply_gst_to_all_cells(wb[sn])

wb.save(path)
print(f'SYS v1.2.0: ChangeLog consolidated, strikethrough rows deleted, Google Sans applied')

# ============================================================
# Summary
# ============================================================
print(f'\nDeleted {len(rows_to_delete)} strikethrough row(s) from SYS 1. ReqSpec')
if len(v120_rows) >= 2:
    print(f'Consolidated {len(v120_rows)} duplicate v1.2.0 ChangeLog entries into 1')
print('All docs updated successfully.')
