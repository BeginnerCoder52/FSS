#!/usr/bin/env python3
"""Fix remaining issues in SYS v1.2.0 after initial generation."""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

DOCS_DIR = os.path.join(os.path.dirname(__file__), '..', 'docs')
SYS_PATH = os.path.join(DOCS_DIR, 'FSS_SystemEngineering_v1.2.0.xlsx')

gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
strike_font = Font(size=11, strike=True, color='666666')
normal_font = Font(size=11)
wrap_align = Alignment(wrap_text=True, vertical='top')

wb = load_workbook(SYS_PATH)

# ===== Fix 1: Modes - remove Relay/LED from cell strings =====
ws = wb['4. Modes']
replacements = {
    6: [  # Observable Behavior (col 6)
        'System control Relay Module and LED Module to state OFF. \n\n',
        'System control Relay Module to state OFF. ',
        'System control Relay to state OFF. ',
    ],
    7: [  # Activity (col 7)
        'System disconnect all Relay control. ',
    ],
    4: [  # Description (col 4)
        'System control Relay Module to state ON.\nSystem control LED Module to state ON (Light Stable 200ms).\n',
    ],
    5: [  # Trigger (col 5)
        'System control Relay Module to ON.\nSystem control LED to ON.\n',
    ],
}
for row in range(1, ws.max_row + 1):
    id_val = str(ws.cell(row=row, column=2).value or '')
    if not id_val.startswith('SYS.MODE'):
        continue
    for col, patterns in replacements.items():
        cell = ws.cell(row=row, column=col)
        if cell.value:
            val = str(cell.value)
            orig = val
            for p in patterns:
                val = val.replace(p, '')
            if val != orig:
                cell.value = val

# ===== Fix 2: Move ChangeLog entry from row 1 to end =====
ws_cl = wb['ChangeLog']
misplaced = ws_cl.cell(row=1, column=1).value
if misplaced == '1.2.0':
    entry = [ws_cl.cell(row=1, column=c).value for c in range(1, 9)]
    for c in range(1, 9):
        ws_cl.cell(row=1, column=c).value = None
    last_row = 1
    for r in range(2, ws_cl.max_row + 1):
        if ws_cl.cell(row=r, column=1).value is not None:
            last_row = r
    insert_at = last_row + 1
    for c, v in enumerate(entry, 1):
        cell = ws_cl.cell(row=insert_at, column=c)
        cell.value = v
        cell.font = normal_font
        cell.alignment = wrap_align

# ===== Fix 3: Mark Light requirements as DEPRECATED in SysReqSpec 15-51 =====
ws_srs = wb['3. SysReqSpec 15-51 v0.3.0']
light_rows = set()
for row in range(1, ws_srs.max_row + 1):
    for col in range(1, ws_srs.max_column + 1):
        v = str(ws_srs.cell(row=row, column=col).value or '')
        if 'SYS.FUCNC.LIGHT' in v or 'SYS.FUNC.LIGHT' in v:
            light_rows.add(row)
            break

for r in light_rows:
    cell = ws_srs.cell(row=r, column=3)  # TÓM TẮT column
    if cell.value:
        cell.value = '[DEPRECATED] ' + str(cell.value)
    # Gray out the row
    for c in range(1, ws_srs.max_column + 1):
        ws_srs.cell(row=r, column=c).fill = gray_fill

# Also mark the LIGHT.03 reference note
for row in range(1, ws_srs.max_row + 1):
    for col in range(1, ws_srs.max_column + 1):
        v = str(ws_srs.cell(row=row, column=col).value or '')
        if 'LIGHT.03' in v:
            ws_srs.cell(row=row, column=col).value = '[DEPRECATED - HW scope change] ' + v

# ===== Fix 4: Mark Light/Relay rows as DEPRECATED in SysArch =====
ws_sa = wb['4. SysArch 04-06 v0.1.0']
for row in range(1, ws_sa.max_row + 1):
    name = str(ws_sa.cell(row=row, column=3).value or '')
    if name in ('Lighting Module', 'Module Relay', 'Lighting Wire', 'LightManager',
                'SensorManager', 'ImageAcquistionController', 'Camera Module'):
        # Prefix name with [DEPRECATED]
        ws_sa.cell(row=row, column=3).value = f'[DEPRECATED] {name}'
        for c in range(1, ws_sa.max_column + 1):
            ws_sa.cell(row=row, column=c).fill = gray_fill
            ws_sa.cell(row=row, column=c).font = Font(size=11, color='999999')

    trace = str(ws_sa.cell(row=row, column=2).value or '')
    if trace == 'SYS.FUCNC.LIGHT.02':
        # Camera Module row
        ws_sa.cell(row=row, column=3).value = '[DEPRECATED] Camera Module'
        for c in range(1, ws_sa.max_column + 1):
            ws_sa.cell(row=row, column=c).fill = gray_fill

# ===== Fix 5: Also clean "1. ReqSpec 17-00" sheet =====
ws_r1 = wb['1. ReqSpec 17-00']
for row in range(1, ws_r1.max_row + 1):
    for col in range(1, ws_r1.max_column + 1):
        v = str(ws_r1.cell(row=row, column=col).value or '')
        if 'SYS.FUCNC.LIGHT' in v or 'SYS.FUNC.LIGHT' in v:
            ws_r1.cell(row=row, column=2).value = '[DEPRECATED] ' + v if ws_r1.cell(row=row, column=2).value else None
            for c in range(1, ws_r1.max_column + 1):
                ws_r1.cell(row=row, column=c).fill = gray_fill
            break
        if 'Relay' in v and 'Module' in v:
            for c in range(1, ws_r1.max_column + 1):
                cval = str(ws_r1.cell(row=row, column=c).value or '')
                if 'Relay' in cval:
                    ws_r1.cell(row=row, column=c).value = '[DEPRECATED - HW scope change] ' + cval

wb.save(SYS_PATH)
print(f'Fixed {SYS_PATH}')
print(' - Modes: removed Relay/LED strings')
print(' - ChangeLog: entry moved to end')
print(' - SysReqSpec: Light rows marked DEPRECATED')
print(' - SysArch: Light/Relay rows marked DEPRECATED')
print(' - ReqSpec 17-00: Light/Relay rows marked DEPRECATED')
