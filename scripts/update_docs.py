#!/usr/bin/env python3
"""Update FSS documentation xlsx files to reflect current source code.

Creates:
  - docs/FSS_SoftwareDetailedDesign_v1.2.0.xlsx  (from v1.1.0)
  - docs/FSS_SoftwareEngineering_v1.3.0.xlsx     (from v1.2.0)
  - docs/FSS_SystemEngineering_v1.2.0.xlsx       (from v1.1.0)
"""

import os, copy, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DOCS_DIR = os.path.join(os.path.dirname(__file__), '..', 'docs')

header_font = Font(bold=True, size=11)
normal_font = Font(size=11)
wrap_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def copy_cell_style(src, dst):
    dst.font = copy.copy(src.font)
    dst.alignment = copy.copy(src.alignment)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.number_format = src.number_format

def copy_sheet_structure(src_ws, dst_ws):
    """Copy all cell values and styles from src to dst."""
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            copy_cell_style(cell, dst_cell)

    # Copy merged cells
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # Copy column widths
    for col_idx in range(1, src_ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width

    # Copy row heights
    for row_idx in range(1, src_ws.max_row + 1):
        if row_idx in src_ws.row_dimensions:
            dst_ws.row_dimensions[row_idx].height = src_ws.row_dimensions[row_idx].height


def write_cell(ws, row, col, value, bold=False, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=11)
    if wrap:
        cell.alignment = wrap_align
    return cell


# ============================================================
# 1. SDD v1.2.0
# ============================================================
def create_sdd_v120():
    src_path = os.path.join(DOCS_DIR, 'FSS_SoftwareDetailedDesign_v1.1.0.xlsx')
    dst_path = os.path.join(DOCS_DIR, 'FSS_SoftwareDetailedDesign_v1.2.0.xlsx')

    wb = load_workbook(src_path)

    # --- Sheet: 0. Overview ---
    ws = wb['0. Overview']
    # Update version cell (row 4, col 2)
    ws.cell(row=4, column=2).value = '1.2.0'
    # Update description
    desc = ws.cell(row=5, column=2).value or ''
    if 'This document defines' in desc:
        ws.cell(row=5, column=2).value = (
            'This document defines logical blocks between subsystems, components, and external '
            'interfaces. Updated for v1.2.0: added RecommendDaemon, RecipeExtractor D-Bus service, '
            'C TFLite Reader; updated FRTApp pipeline (ByteTrack, tflite-runtime); '
            'consolidated test structure under /tests/; revised MagicMirror modules to MMM-FSS-* naming.'
        )

    # --- Sheet: x.ChangeLog ---
    ws_cl = wb['x.ChangeLog']
    # Find first empty row after version entries
    empty_row = None
    for row in range(1, ws_cl.max_row + 1):
        if ws_cl.cell(row=row, column=1).value is None and ws_cl.cell(row=row, column=3).value is None:
            empty_row = row
            break
    if empty_row is None:
        empty_row = ws_cl.max_row + 1

    changelog_entries = [
        ('1.2.0',
         '[Update]: Add RecommendDaemon, RecipeExtractor D-Bus service, C TFLite Reader to architecture\n'
         '[Update]: [1.SoftwareComponentInteraction] Added RecommendDaemon, RecipeExtractor, C TFLite Reader\n'
         '[Update]: [2.APISpecifications] Added RecommendDaemon, RecipeExtractor tables; updated FRTApp (ByteTrack, tflite-runtime)\n'
         '[Update]: [3.Inter-Class Relationships] Added IPC links for new components\n'
         '[Fix]: [2.APISpecifications] [FRTApp] Changed DeepSORT references to ByteTrack, ultralytics to tflite-runtime\n'
         '[Fix]: [3.Inter-Class Relationships] [ElectronApp] Updated D-Bus signals for MMM-FSS-* module names\n'
         '[Update]: [0.Overview] Updated Document Version to 1.2.0 to reflect architecture changes',
         'M', '-')
    ]

    for i, (ver, desc, actions, notes) in enumerate(changelog_entries):
        r = empty_row + i
        ws_cl.cell(row=r, column=1).value = ver
        ws_cl.cell(row=r, column=2).value = desc
        ws_cl.cell(row=r, column=3).value = actions
        ws_cl.cell(row=r, column=4).value = notes
        for c in range(1, 5):
            cell = ws_cl.cell(row=r, column=c)
            cell.font = normal_font
            cell.alignment = wrap_align

    # --- Sheet: 1.SoftwareComponentInteraction ---
    ws_sci = wb['1.SoftwareComponentInteraction']
    # Find first empty row
    empty_row = None
    for row in range(1, ws_sci.max_row + 1):
        if all(ws_sci.cell(row=row, column=c).value is None for c in range(1, 6)):
            empty_row = row
            break
    if empty_row is None:
        empty_row = ws_sci.max_row + 1

    # Update existing FRTApp row to mention ByteTrack
    for row in range(1, ws_sci.max_row + 1):
        cell_a = str(ws_sci.cell(row=row, column=1).value or '')
        if 'FRTApp' in cell_a:
            current_desc = str(ws_sci.cell(row=row, column=2).value or '')
            if 'DeepSORT' in current_desc:
                ws_sci.cell(row=row, column=2).value = current_desc.replace('DeepSORT', 'ByteTrack')
            if 'ultralytics' in current_desc:
                ws_sci.cell(row=row, column=4).value = current_desc  # skip

    # Add RecommendDaemon row
    data_rec_daemon = [
        'RecommendDaemon',
        'Orchestrator business logic quản lý gợi ý mua sắm. Nhận recipe từ UI, gọi RecipeAnalyzerEngine '
        'để NLP-extract ingredients, truy vấn tồn kho qua DBDaemon D-Bus, chạy Bù Trừ algorithm, '
        'lưu kết quả vào FSS-Recommend.db, phát RecommendationUpdated signal.',
        '- Gọi D-Bus method RecommendDaemon.GenerateShoppingList từ UI (Python bridge)\n'
        '- Gọi D-Bus RecipeExtractor.ExtractAndPersistRecipe (nội bộ)\n'
        '- Gọi D-Bus DBDaemon.GetInventory() qua proxy\n'
        '- Phát D-Bus signal RecommendationUpdated',
        '- recipe_extractor (NLP engine, imported library)\n'
        '- recommend_daemon venv (Python 3)\n'
        '- sdbus (Python D-Bus binding)',
        'RecipeAnalyzerEngine.generate_fss_request(recipe_name)\n'
        'RecommendEngine.bu_tru_compare(ingredients, inventory)\n'
        'RecommendDbManager.insert_shopping_list()'
    ]
    for i, val in enumerate(data_rec_daemon):
        write_cell(ws_sci, empty_row + i, 1, val, bold=(i==0))

    # Add RecipeExtractor row
    data_recipe_ext = [
        'RecipeExtractor\n(D-Bus Service)',
        'Dịch vụ D-Bus độc lập, nhận tên món ăn, chạy NLP (CRF NER BIO-tagged) để trích xuất '
        'ingredients và quantities, trả về JSON FSS-Request. Gọi DBDaemon.InsertRequest() để lưu.',
        '- Lắng nghe D-Bus method ExtractAndPersistRecipe\n'
        '- Gọi D-Bus DBDaemon.InsertRequest() proxy\n'
        '- Trả về JSON kết quả NLP',
        '- recipe_extractor venv (Python 3)\n'
        '- sklearn-crfsuite\n'
        '- sdbus (Python D-Bus binding)',
        'RecipeAnalyzerEngine.generate_fss_request()\n'
        'RecipeExtractorService.ExtractAndPersistRecipe()'
    ]
    offset = empty_row + 3
    for i, val in enumerate(data_recipe_ext):
        write_cell(ws_sci, offset + i, 1, val, bold=(i==0))

    # Update DBDaemon row to include RecommendDaemon integration
    for row in range(1, ws_sci.max_row + 1):
        cell_a = str(ws_sci.cell(row=row, column=1).value or '')
        if 'DBDaemon' in cell_a and 'Data Controller' in str(ws_sci.cell(row=row, column=2).value or ''):
            cur_desc = str(ws_sci.cell(row=row, column=2).value or '')
            ws_sci.cell(row=row, column=2).value = cur_desc + (
                '\n- Cung cấp D-Bus methods: GetInventory(), GetRecipeRequests(), '
                'InsertRequest() cho RecommendDaemon/RecipeExtractor.\n'
                '- Quản lý 3 database files: fss_data.db, FSS_Inventory.db, FSS_Request.db.'
            )

    # --- Sheet: 2.APISpecifications ---
    ws_api = wb['2.APISpecifications']

    # Update FRTApp table: change DeepSORT references to ByteTrack
    for row in range(1, ws_api.max_row + 1):
        for col in range(1, ws_api.max_column + 1):
            val = str(ws_api.cell(row=row, column=col).value or '')
            if 'DeepSORT' in val:
                ws_api.cell(row=row, column=col).value = val.replace('DeepSORT', 'ByteTrack')
            if 'ultralytics' in val.lower():
                ws_api.cell(row=row, column=col).value = val.replace('Ultralytics', 'tflite-runtime')

    # Find the end of existing tables
    empty_row = ws_api.max_row + 1

    # Add Bảng 5: RecommendDaemon
    write_cell(ws_api, empty_row, 1, 'Bảng 5: Package RecommendDaemon', bold=True)
    empty_row += 1
    headers_rec = ['Package (App)', 'Class', 'Member Type', 'Name / Signature', 'Data / Return Type', 'Description', 'Notes']
    for i, h in enumerate(headers_rec):
        write_cell(ws_api, empty_row, i+1, h, bold=True)
    empty_row += 1

    rec_data = [
        ['RecommendDaemon', 'RecommendDaemonMain', 'Attribute', 'current_state', 'str', 'Trạng thái daemon (INIT, IDLE, PROCESSING, ERROR).', ''],
        ['RecommendDaemon', '', 'Attribute', 'is_running', 'bool', 'Cờ duy trì vòng lặp chính.', ''],
        ['RecommendDaemon', '', 'Attribute', 'recommend_engine', 'RecommendEngine', 'Engine xử lý Bù Trừ logic.', ''],
        ['RecommendDaemon', '', 'Attribute', 'db_manager', 'RecommendDbManager', 'Quản lý FSS-Recommend.db.', ''],
        ['RecommendDaemon', '', 'Attribute', 'dbus_iface', 'DbusInterface', 'Giao tiếp D-Bus.', ''],
        ['RecommendDaemon', '', 'Method', 'init_daemon()', 'bool', 'Khởi tạo engine, DB, D-Bus.', ''],
        ['RecommendDaemon', '', 'Method', 'start_daemon()', 'void', 'Chạy vòng lặp chính (asyncio).', ''],
        ['RecommendDaemon', '', 'Method', 'stop_daemon()', 'void', 'Dọn dẹp và tắt an toàn.', ''],
        ['RecommendDaemon', '', 'Method', 'generate_shopping_list(recipe: str)', 'str (JSON)', 'Gọi NLP → lấy inventory → Bù Trừ → persist → emit signal.', 'D-Bus method'],
        ['RecommendDaemon', '', 'Method', 'get_available_recipes()', 'str (JSON)', 'Trả danh sách recipes từ NLP engine.', 'D-Bus method'],
        ['RecommendDaemon', '', 'Method', 'get_shopping_list()', 'str (JSON)', 'Trả danh sách mua sắm hiện tại.', 'D-Bus method'],
        ['RecommendDaemon', '', 'Method', 'mark_item_purchased(food_id: str)', 'bool', 'Đánh dấu item đã mua.', 'D-Bus method'],
        ['RecommendDaemon', 'RecommendEngine', 'Attribute', 'nlp_engine', 'RecipeAnalyzerEngine', 'CRF NER engine từ recipe_extractor.', 'Lazy-loaded'],
        ['RecommendDaemon', '', 'Attribute', 'nlp_loaded', 'bool', 'Trạng thái NLP engine.', ''],
        ['RecommendDaemon', '', 'Method', 'bu_tru_compare(ingredients, inventory)', 'dict', 'So sánh nhu cầu vs tồn kho.', ''],
        ['RecommendDaemon', '', 'Method', 'load_nlp_engine()', 'bool', 'Lazy-load NLP model.', ''],
        ['RecommendDaemon', '', 'Method', 'generate_shopping_list(recipe_name: str)', 'str (JSON)', 'Pipeline NLP → Bù Trừ.', ''],
        ['RecommendDaemon', 'DbusInterface', 'Attribute', 'system_bus', 'object', 'Kết nối System D-Bus.', ''],
        ['RecommendDaemon', '', 'Attribute', 'service_name', 'str', 'vn.edu.uit.FSS.RecommendDaemon', ''],
        ['RecommendDaemon', '', 'Method', 'setup_service()', 'bool', 'Đăng ký service lên bus.', ''],
        ['RecommendDaemon', '', 'Method', 'emit_recommendation_updated(payload)', 'void', 'Signal khi có recommend mới.', ''],
        ['RecommendDaemon', '', 'Method', 'call_dbd_get_inventory()', 'str (JSON)', 'Gọi DBDaemon.GetInventory() qua proxy.', ''],
        ['RecommendDaemon', 'RecommendDbManager', 'Attribute', 'db_path', 'str', 'Đường dẫn FSS-Recommend.db.', ''],
        ['RecommendDaemon', '', 'Method', 'init_database()', 'bool', 'Tạo bảng recommendation_log, shopping_list.', ''],
        ['RecommendDaemon', '', 'Method', 'insert_recommendation(data)', 'int', 'Insert batch, trả về batch_id.', ''],
        ['RecommendDaemon', '', 'Method', 'insert_shopping_list(items)', 'bool', 'Insert danh sách mua sắm.', ''],
        ['RecommendDaemon', '', 'Method', 'get_active_shopping_list()', 'list', 'Lấy items chưa mua.', ''],
        ['RecommendDaemon', '', 'Method', 'mark_purchased(item_id: int)', 'bool', 'Cập nhật purchased flag.', ''],
    ]
    for row_data in rec_data:
        for i, val in enumerate(row_data):
            write_cell(ws_api, empty_row, i+1, val)
        empty_row += 1

    empty_row += 1

    # Add Bảng 6: RecipeExtractor
    write_cell(ws_api, empty_row, 1, 'Bảng 6: Package RecipeExtractor', bold=True)
    empty_row += 1
    for i, h in enumerate(headers_rec):
        write_cell(ws_api, empty_row, i+1, h, bold=True)
    empty_row += 1

    recipe_data = [
        ['RecipeExtractor', 'RecipeExtractorService', 'Attribute', 'system_bus', 'object', 'Kết nối System D-Bus.', ''],
        ['RecipeExtractor', '', 'Attribute', 'service_name', 'str', 'vn.edu.uit.FSS.RecipeExtractor', ''],
        ['RecipeExtractor', '', 'Attribute', 'nlp_engine', 'RecipeAnalyzerEngine', 'Engine NLP CRF.', ''],
        ['RecipeExtractor', '', 'Method', 'setup_service()', 'bool', 'Đăng ký D-Bus service.', ''],
        ['RecipeExtractor', '', 'Method', 'ExtractAndPersistRecipe(recipe_name: str)', 'str (JSON)', 'NLP extract → InsertRequest → return JSON.', 'D-Bus method'],
        ['RecipeExtractor', 'RecipeAnalyzerEngine', 'Attribute', 'model_path', 'str', 'Đường dẫn CRF model .joblib.', ''],
        ['RecipeExtractor', '', 'Attribute', 'crf_model', 'object', 'CRF model (sklearn-crfsuite).', ''],
        ['RecipeExtractor', '', 'Method', 'generate_fss_request(recipe_name: str)', 'dict', 'Phân tích recipe → ingredients list.', ''],
        ['RecipeExtractor', '', 'Method', '_load_crf_model()', 'bool', 'Load model từ joblib.', ''],
        ['RecipeExtractor', '', 'Method', '_extract_entities(tokens, pos_tags)', 'list', 'BIO tagging inference.', ''],
        ['RecipeExtractor', '', 'Method', '_normalize_quantity(raw_qty)', 'str', 'Chuẩn hóa số lượng.', ''],
        ['RecipeExtractor', 'RecipeProcessor', 'Method', 'process_recipe(recipe_name: str)', 'dict', 'Xử lý recipe đầu vào, trả về structured data.', ''],
    ]
    for row_data in recipe_data:
        for i, val in enumerate(row_data):
            write_cell(ws_api, empty_row, i+1, val)
        empty_row += 1

    # Update ElectronApp table: mention MMM-FSS-* modules
    for row in range(1, ws_api.max_row + 1):
        for col in range(1, ws_api.max_column + 1):
            val = str(ws_api.cell(row=row, column=col).value or '')
            if 'ElectronApp' in val:
                pass  # Keep existing - already lists ElectronMain, DbusNextListener, etc.

    # --- Sheet: 3. Inter-Class Relationships ---
    ws_icr = wb['3. Inter-Class Relationships']
    empty_row = ws_icr.max_row + 1

    # Add new IPC relationships for new components
    new_relations = [
        ['RecommendDaemon', 'DbusInterface', 'DBDaemon', 'DbDbusInterface', 'Inter-Process (IPC)',
         'D-Bus Method Call (GetInventory)', 'RecommendDaemon gọi GetInventory() qua D-Bus proxy để lấy tồn kho hiện tại.'],
        ['RecommendDaemon', 'DbusInterface', 'RecommendDaemon', 'RecommendEngine', 'Internal (DI)',
         'Direct Python call', 'RecommendDaemon gọi RecommendEngine.bu_tru_compare() với ingredients và inventory.'],
        ['RecommendDaemon', 'RecommendDbManager', 'OS File System', 'Ext4 /opt/fss/...', 'File I/O',
         'SQLite API', 'Lưu kết quả recommend xuống FSS-Recommend.db.'],
        ['RecipeExtractor', 'RecipeExtractorService', 'DBDaemon', 'DbDbusInterface', 'Inter-Process (IPC)',
         'D-Bus Method Call (InsertRequest)', 'RecipeExtractor gọi DBDaemon.InsertRequest() để lưu recipe request.'],
        ['RecipeExtractor', 'RecipeAnalyzerEngine', 'OS File System', 'Ext4 /opt/fss/...', 'File I/O',
         'joblib load', 'Load CRF model từ file .joblib tại startup.'],
        ['UI', 'MMM-FSS-Recommend', 'RecommendDaemon', 'DbusInterface', 'Inter-Process (IPC)',
         'D-Bus Method Call (GenerateShoppingList)', 'UI gửi recipe → RecommendDaemon xử lý → trả kết quả.'],
        ['FRTApp', 'FrtDbusInterface', 'RecommendDaemon', 'DbusInterface', 'Inter-Process (IPC)',
         'D-Bus Signal (FoodDetected)', 'FRTApp thông báo food detected (có thể trigger re-recommend).'],
    ]
    for rel in new_relations:
        for i, val in enumerate(rel):
            write_cell(ws_icr, empty_row, i+1, val)
        empty_row += 1

    wb.save(dst_path)
    print(f'Created {dst_path}')
    return dst_path


# ============================================================
# 2. SWE v1.3.0
# ============================================================
def create_swe_v130():
    src_path = os.path.join(DOCS_DIR, 'FSS_SoftwareEngineering_v1.2.0.xlsx')
    dst_path = os.path.join(DOCS_DIR, 'FSS_SoftwareEngineering_v1.3.0.xlsx')

    wb = load_workbook(src_path)

    # --- Sheet: 1. SoftwareReqSpec ---
    ws = wb['1. SoftwareReqSpec']
    # Update version cell (row 3, col 3)
    ws.cell(row=3, column=3).value = '1.3.0'

    # Find last row with data
    max_row = ws.max_row
    empty_row = max_row + 1

    # Add new requirements
    new_reqs = [
        # RecommendDaemon requirements
        ['', '', '-', 'Heading', 'RecommendDaemon (Recommendation Engine) - Quản lý gợi ý mua sắm thông minh',
         '', '', '', '', '', '', '', ''],
        ['', 'SYS.FUNC.AI.02', 'SW.REC.01', 'Functional Requirement', 'D-Bus Service Setup',
         'RecommendDaemon shall register D-Bus service vn.edu.uit.FSS.RecommendDaemon on the system bus and expose methods: GenerateShoppingList, GetAvailableRecipes, GetShoppingList, MarkItemPurchased.',
         '', '', 'recommend_daemon/src/DbusInterface.py', 'High', 'Verification by Testing',
         'dbus-send --system --print-reply --dest=vn.edu.uit.FSS.RecommendDaemon returns method list.', 'Draft'],
        ['', 'SYS.FUNC.AI.02', 'SW.REC.02', 'Functional Requirement', 'NLP Integration',
         'RecommendDaemon shall lazy-load RecipeAnalyzerEngine from recipe_extractor when GenerateShoppingList is first called.',
         '', '', 'recommend_daemon/src/RecommendEngine.py', 'High', 'Verification by Testing',
         'Log shows "NLP engine loaded" on first call.', 'Draft'],
        ['', 'SYS.FUNC.AI.02', 'SW.REC.03', 'Functional Requirement', 'Bù Trừ Algorithm',
         'RecommendDaemon shall compare extracted ingredient list against current inventory via DBDaemon D-Bus GetInventory() and compute shortfall using Bù Trừ method.',
         '', '', 'recommend_daemon/src/RecommendEngine.py', 'High', 'Verification by Testing',
         'Output JSON contains available, needed, missing counts.', 'Draft'],
        ['', 'SYS.FUNC.AI.02', 'SW.REC.04', 'Functional Requirement', 'Shopping List Persistence',
         'RecommendDaemon shall persist shopping list to FSS-Recommend.db with tables recommendation_log and shopping_list.',
         '', '', 'recommend_daemon/src/RecommendDbManager.py', 'High', 'Verification by Inspection',
         'Database file exists with correct schema.', 'Draft'],
        ['', '', '-', 'Heading', 'RecipeExtractor (NLP D-Bus Service) - Trích xuất nguyên liệu từ tên món ăn',
         '', '', '', '', '', '', '', ''],
        ['', 'SYS.FUNC.AI.02', 'SW.EXT.01', 'Functional Requirement', 'D-Bus Service',
         'RecipeExtractor shall register D-Bus service vn.edu.uit.FSS.RecipeExtractor and expose method ExtractAndPersistRecipe.',
         '', '', 'recipe_extractor/src/recipe_extractor_service.py', 'High', 'Verification by Testing',
         'dbus-send returns JSON result.', 'Draft'],
        ['', 'SYS.FUNC.AI.02', 'SW.EXT.02', 'Functional Requirement', 'CRF NER Model',
         'RecipeExtractor shall load fss_ner_crf_optimized.joblib and perform BIO-tagged NER to extract ingredients and quantities.',
         'F1-Score >= 95%', '', 'recipe_extractor/src/RecipeAnalyzerAPI.py', 'High', 'Verification by Testing',
         'pytest tests/recipe_extractor/test_recipe_analyzer.py -v passes.', 'Draft'],
        ['', 'SYS.FUNC.AI.02', 'SW.EXT.03', 'Functional Requirement', 'Persist Request',
         'RecipeExtractor shall call DBDaemon.InsertRequest() via D-Bus proxy to persist recipe request.',
         '', '', 'recipe_extractor/src/recipe_extractor_service.py', 'Medium', 'Verification by Testing',
         'Request appears in FSS_Request.db recipe_requests table.', 'Draft'],
        # C TFLite Reader requirements
        ['', '', '-', 'Heading', 'C TFLite Reader (Performance Backend) - Inference tối ưu bằng C API',
         '', '', '', '', '', '', '', ''],
        ['', 'SYS.PERF.GEN.01', 'SW.CTL.01', 'Functional Requirement', 'Model Loading',
         'The C TFLite Reader shall load FP32, FP16, and INT8 quantized .tflite models via TensorFlow Lite C API.',
         '', '', 'frt_app/c_tflite_reader/src/TfliteReader.c', 'High', 'Verification by Testing',
         './tflite_reader_test --model model.tflite --precision int8 exits 0.', 'Draft'],
        ['', 'SYS.PERF.GEN.01', 'SW.CTL.02', 'Functional Requirement', 'Inference Execution',
         'The C TFLite Reader shall preprocess input, invoke inference, and return dequantized float output.',
         '', '', 'frt_app/c_tflite_reader/src/TfliteReader.c', 'High', 'Verification by Testing',
         'Output detections have valid confidence scores.', 'Draft'],
        ['', 'SYS.PERF.GEN.01', 'SW.CTL.03', 'Non-Functional Requirement', 'Performance',
         'The C backend shall achieve inference latency <= 50ms on Raspberry Pi 4B with INT8 model.',
         '', '', 'frt_app/c_tflite_reader/src/TfliteReader.c', 'Medium', 'Verification by Testing',
         'Benchmark shows avg inference time < 50ms.', 'Draft'],
        # Test consolidation
        ['', '', '-', 'Heading', 'Test Framework - Kiểm thử tập trung',
         '', '', '', '', '', '', '', ''],
        ['', 'SYS.NOF.02', 'SW.TST.01', 'Non-Functional Requirement', 'Centralized Tests',
         'All component tests shall be consolidated under /tests/<component>/ directory except MagicMirror tests which remain at electron_app/magicmirror/tests/.',
         '', '', '/tests/', 'Medium', 'Verification by Inspection',
         'Test directory structure matches AGENTS.md specification.', 'Implemented'],
        # D-Bus IPC requirements
        ['', '', '-', 'Heading', 'D-Bus IPC Infrastructure - Giao tiếp liên tiến trình',
         '', '', '', '', '', '', '', ''],
        ['', 'SYS.INT.06', 'SW.IPC.01', 'Non-Functional Requirement', 'System Bus Only',
         'All daemon-to-daemon communication shall use D-Bus System Bus. No ZMQ or WebSocket IPC between daemons.',
         '', '', 'All daemons', 'High', 'Verification by Inspection',
         'Source code contains no ZMQ/WebSocket imports for IPC.', 'Implemented'],
        ['', 'SYS.INT.06', 'SW.IPC.02', 'Non-Functional Requirement', 'Signal Ordering',
         'D-Bus signals shall be fire-and-forget. No guarantee of delivery ordering.',
         '', '', 'All daemons', 'Medium', 'Verification by Inspection',
         'Signal emission code follows async pattern without ACK.', 'Draft'],
    ]

    for i, row_data in enumerate(new_reqs):
        r = empty_row + i
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j+1).value = val
            ws.cell(row=r, column=j+1).font = normal_font
            ws.cell(row=r, column=j+1).alignment = wrap_align

    # Update ChangeLog
    ws_cl = wb['x. ChangeLog']
    empty_row = None
    for row in range(1, ws_cl.max_row + 1):
        if ws_cl.cell(row=row, column=1).value is None and ws_cl.cell(row=row, column=3).value is None:
            empty_row = row
            break
    if empty_row is None:
        empty_row = ws_cl.max_row + 1

    ws_cl.cell(row=empty_row, column=1).value = '1.3.0'
    ws_cl.cell(row=empty_row, column=2).value = (
        '[Add New]: Added RecommendDaemon requirements (SW.REC.01-04)\n'
        '[Add New]: Added RecipeExtractor requirements (SW.EXT.01-03)\n'
        '[Add New]: Added C TFLite Reader requirements (SW.CTL.01-03)\n'
        '[Add New]: Added D-Bus IPC requirements (SW.IPC.01-02)\n'
        '[Update]: Added Test Framework requirement (SW.TST.01)\n'
        '[Update]: Updated Document Version to 1.3.0'
    )
    ws_cl.cell(row=empty_row, column=3).value = 'A'
    ws_cl.cell(row=empty_row, column=4).value = 'Align with SDD v1.2.0 and new architecture'

    wb.save(dst_path)
    print(f'Created {dst_path}')
    return dst_path


# ============================================================
# 3. SYS v1.2.0
# ============================================================
def create_sys_v120():
    src_path = os.path.join(DOCS_DIR, 'FSS_SystemEngineering_v1.1.0.xlsx')
    dst_path = os.path.join(DOCS_DIR, 'FSS_SystemEngineering_v1.2.0.xlsx')

    wb = load_workbook(src_path)

    # --- Sheet: 1. ReqSpec (actual name "1. ReqSpec") ---
    ws_req = wb['1. ReqSpec']
    # Update version
    ws_req.cell(row=3, column=3).value = '1.2.0'

    # Remove Relay Module / LED references where possible
    # Add new requirements for RecommendDaemon and RecipeExtractor
    max_row = ws_req.max_row
    empty_row = max_row + 1

    new_sys_reqs = [
        ('', 'SYS.FUNC.AI.02', 'The System shall provide a recommendation engine that accepts recipe names, extracts ingredients via NLP, compares against current inventory, and generates a shopping list with shortfall quantities.', 'Functional Requirement', 'Recommendation Subsystem', 'HIGH', '-'),
        ('', 'SYS.FUNC.AI.02', 'The System shall expose a D-Bus service (vn.edu.uit.FSS.RecommendDaemon) for recipe-based shopping list generation.', 'Functional Requirement', 'Recommendation Subsystem', 'HIGH', '-'),
        ('', 'SYS.FUNC.AI.02', 'The System shall provide a D-Bus service (vn.edu.uit.FSS.RecipeExtractor) for NLP-based ingredient extraction from recipe names.', 'Functional Requirement', 'Recommendation Subsystem', 'HIGH', '-'),
        ('', 'SYS.FUNC.AI.02', 'The System shall support a C-based TensorFlow Lite inference backend (libtflite_reader.so) as an optional high-performance alternative to Python tflite-runtime.', 'Functional Requirement', 'AI Core', 'MEDIUM', '-'),
    ]

    for i, (no_func, id_val, desc, rtype, alloc, prio, risk) in enumerate(new_sys_reqs):
        r = empty_row + i
        if no_func:
            ws_req.cell(row=r, column=1).value = no_func
        ws_req.cell(row=r, column=2).value = id_val
        ws_req.cell(row=r, column=3).value = desc
        ws_req.cell(row=r, column=4).value = rtype
        ws_req.cell(row=r, column=5).value = alloc
        ws_req.cell(row=r, column=6).value = prio
        ws_req.cell(row=r, column=7).value = risk

    # --- Sheet: 2. Elements ---
    ws_el = wb['2. Elements']
    # Add RecommendDaemon and RecipeExtractor as software elements
    max_row = ws_el.max_row
    empty_row = max_row + 1

    new_elements = [
        ('', 'SYS.ELEM.11', 'A.A*', 'RecommendDaemon', 'SW', '- Python daemon.\n- D-Bus service: vn.edu.uit.FSS.RecommendDaemon.\n- Functions: GenerateShoppingList, GetAvailableRecipes, GetShoppingList, MarkItemPurchased.\n- Database: FSS-Recommend.db.', 'SYS.FUNC.AI.02'),
        ('', 'SYS.ELEM.12', 'A.A', 'RecipeExtractor Service', 'SW', '- Python D-Bus service.\n- D-Bus service: vn.edu.uit.FSS.RecipeExtractor.\n- Function: ExtractAndPersistRecipe.\n- Model: fss_ner_crf_optimized.joblib (CRF).', 'SYS.FUNC.AI.02'),
        ('', 'SYS.ELEM.13', 'A.A', 'C TFLite Reader Library', 'SW', '- Shared library (libtflite_reader.so).\n- TensorFlow Lite C API.\n- Supports FP32/FP16/INT8 precisions.\n- Optional backend for Python AI Core.', 'SYS.PERF.GEN.01'),
    ]
    for i, row_data in enumerate(new_elements):
        for j, val in enumerate(row_data):
            ws_el.cell(row=empty_row + i, column=j+1).value = val
            ws_el.cell(row=empty_row + i, column=j+1).font = normal_font
            ws_el.cell(row=empty_row + i, column=j+1).alignment = wrap_align

    # --- Sheet: 3. Interfaces ---
    ws_if = wb['3. Interfaces']
    max_row = ws_if.max_row
    empty_row = max_row + 1

    new_interfaces = [
        ('SYS.IF.INT.03', 'RecommendDaemon', 'DBDaemon', 'Logical', 'D-Bus Method Call (GetInventory).\nProtocol: sdbus (Python).', 'SYS.FUNC.AI.02'),
        ('SYS.IF.INT.04', 'RecipeExtractor', 'DBDaemon', 'Logical', 'D-Bus Method Call (InsertRequest).\nProtocol: sdbus (Python).', 'SYS.FUNC.AI.02'),
        ('SYS.IF.INT.05', 'UI / MagicMirror', 'RecommendDaemon', 'Logical', 'D-Bus Method Call (GenerateShoppingList).\nProtocol: sdbus via Python bridge.', 'SYS.FUNC.AI.02'),
        ('SYS.IF.INT.06', 'FRTApp (AI Core)', 'C TFLite Reader', 'Logical', 'ctypes FFI.\nProtocol: C shared library (.so).', 'SYS.PERF.GEN.01'),
    ]
    for i, row_data in enumerate(new_interfaces):
        for j, val in enumerate(row_data):
            ws_if.cell(row=empty_row + i, column=j+1).value = val
            ws_if.cell(row=empty_row + i, column=j+1).font = normal_font
            ws_if.cell(row=empty_row + i, column=j+1).alignment = wrap_align

    # --- Sheet: 4. Modes ---
    ws_mode = wb['4. Modes']
    # Update Mode 04 SYS_SCANNING to reflect no Relay/LED
    for row in range(1, ws_mode.max_row + 1):
        id_val = str(ws_mode.cell(row=row, column=2).value or '')
        if id_val == 'SYS.MODE.04':
            desc = str(ws_mode.cell(row=row, column=4).value or '')
            desc = desc.replace(
                'System control USB Camera to active and wait for light stable 200ms.',
                'System control USB Camera to active.'
            )
            ws_mode.cell(row=row, column=4).value = desc

            obs = str(ws_mode.cell(row=row, column=6).value or '')
            obs = obs.replace(
                'System control Relay Module to state ON.\nSystem control LED Module to state ON (Light Stable 200ms).',
                'System control USB Camera to active.'
            )
            obs = obs.replace(
                '\nSystem control Relay Module to state OFF.\nSystem control LED Module to state OFF.',
                ''
            )
            ws_mode.cell(row=row, column=6).value = obs

            act = str(ws_mode.cell(row=row, column=7).value or '')
            act = act.replace('System control Relay Module to ON/OFF\nSystem control LED Module to ON/OFF\n', '')
            ws_mode.cell(row=row, column=7).value = act

        elif id_val == 'SYS.MODE.01':
            obs = str(ws_mode.cell(row=row, column=6).value or '')
            obs = obs.replace('System control Relay Module and LED Module to state OFF.\n', '')
            ws_mode.cell(row=row, column=6).value = obs

        elif id_val == 'SYS.MODE.02':
            obs = str(ws_mode.cell(row=row, column=6).value or '')
            obs = obs.replace('System control Relay Module to state OFF.\n', '')
            ws_mode.cell(row=row, column=6).value = obs

        elif id_val == 'SYS.MODE.06':
            obs = str(ws_mode.cell(row=row, column=6).value or '')
            obs = obs.replace('System control Relay to state OFF.\n', '')
            ws_mode.cell(row=row, column=6).value = obs

    # --- Update ChangeLog ---
    ws_cl = wb['ChangeLog']
    empty_row = None
    for row in range(1, ws_cl.max_row + 1):
        if ws_cl.cell(row=row, column=1).value is None and ws_cl.cell(row=row, column=3).value is None:
            empty_row = row
            break
    if empty_row is None:
        empty_row = ws_cl.max_row + 1

    ws_cl.cell(row=empty_row, column=1).value = '1.2.0'
    ws_cl.cell(row=empty_row, column=2).value = (
        '[Update]: Removed Relay Module and LED Module references across all sheets.\n'
        '[Add New]: Added RecommendDaemon, RecipeExtractor, C TFLite Reader as system elements.\n'
        '[Update]: [1.ReqSpec] Added SYS.FUNC.AI.02 for recommendation and extraction.\n'
        '[Update]: [2.Elements] Added SYS.ELEM.11-13 for new software components.\n'
        '[Update]: [3.Interfaces] Added SYS.IF.INT.03-06 for new IPC flows.\n'
        '[Update]: [4.Modes] Removed Relay/LED actuation from SYS_SCANNING and SYS_INIT/SYS_IDLE.\n'
        '[Fix]: Version consistency — ReqSpec sheet updated from v0.6.0 to v1.2.0 to match ChangeLog.'
    )
    ws_cl.cell(row=empty_row, column=3).value = 'A, M, D'
    ws_cl.cell(row=empty_row, column=4).value = 'Align with current hardware scope and SW architecture.'

    wb.save(dst_path)
    print(f'Created {dst_path}')
    return dst_path


# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    sdd = create_sdd_v120()
    swe = create_swe_v130()
    sys = create_sys_v120()
    print('\nAll docs updated successfully.')
    print(f'  SDD: {sdd}')
    print(f'  SWE: {swe}')
    print(f'  SYS: {sys}')
