#!/usr/bin/env python3
"""Comprehensive formatting fix for all 3 updated FSS xlsx documents.

Fixes:
1. SDD v1.2.0: ChangeLog position/format, API table formatting, new row formatting
2. SWE v1.3.0: ChangeLog entry format, requirement row formatting
3. SYS v1.2.0: ChangeLog entry format, new elements/interfaces formatting
"""

import os, copy, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter

DOCS = os.path.join(os.path.dirname(__file__), '..', 'docs')

# ============================================================
# Common style definitions
# ============================================================

# Google Sans Text font
gst = Font(name='Google Sans Text', size=11, color=Color(theme=1))
gst_bold = Font(name='Google Sans Text', size=11, bold=True, color=Color(theme=1))
gst_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
gst_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Calibri (for ChangeLog)
cal = Font(name='Calibri', size=11, color=Color(theme=1))
cal_bold = Font(name='Calibri', size=11, bold=True, color=Color(theme=1))
cal_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
cal_center_top = Alignment(horizontal='center', vertical='top', wrap_text=True)
cal_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Thin border
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
thin_left = Border(left=thin_side)
thin_left_right = Border(left=thin_side, right=thin_side)

# Table header fill (theme 5, tint 0.8)
header_fill = PatternFill(fill_type='solid', fgColor=Color(theme=5, tint=0.7999816888943144))

# ============================================================
# Helper: copy formatting from template cell
# ============================================================
def apply_style(cell, font, alignment=None, fill=None, border=None):
    cell.font = copy.copy(font)
    if alignment:
        cell.alignment = copy.copy(alignment)
    if fill:
        cell.fill = copy.copy(fill)
    if border:
        cell.border = copy.copy(border)

def style_api_data_row(cell, is_first_col_bold=False):
    """Style a data row in SDD API table (Google Sans Text, thin border)."""
    f = gst_bold if is_first_col_bold else gst
    apply_style(cell, f, border=thin_border)

def style_api_header_row(cell):
    """Style a header row in SDD API table."""
    apply_style(cell, gst_bold, alignment=gst_center, fill=header_fill, border=thin_border)

def style_api_section_header(cell):
    """Style a 'Bảng X: Package Y' section header."""
    apply_style(cell, gst_bold, alignment=gst_center, border=thin_border)

# ============================================================
# Fix SDD v1.2.0
# ============================================================
def fix_sdd():
    path = os.path.join(DOCS, 'FSS_SoftwareDetailedDesign_v1.2.0.xlsx')
    wb = load_workbook(path)

    # ----- Fix 1: ChangeLog -----
    ws = wb['x.ChangeLog']

    # Check if entry at row 1 needs to be moved
    if ws.cell(row=1, column=1).value == '1.2.0':
        # Read the full entry
        desc = str(ws.cell(row=1, column=2).value or '')
        actions = str(ws.cell(row=1, column=3).value or '')
        notes = str(ws.cell(row=1, column=4).value or '')

        # Clear row 1-5 (the misplaced content)
        for r in range(1, 6):
            for c in range(1, 9):
                ws.cell(row=r, column=c).value = None

        # Restore row 1-5 (instructions section of ChangeLog)
        instructions = [
            (1, 6, 'Change Description:', None, None, None),
        ]
        for c in range(6, 9):
            ws.cell(row=1, column=c).value = instructions[0][c-6] if c == 6 else None
            apply_style(ws.cell(row=1, column=c), cal, alignment=cal_center_top)

        ws.cell(row=2, column=7).value = 'Must have [Reason] [Sheet changed][Sections changed][Details]'
        apply_style(ws.cell(row=2, column=7), cal, alignment=cal_center_top)

        ws.cell(row=3, column=7).value = 'Reason'
        apply_style(ws.cell(row=3, column=7), cal_bold, alignment=cal_center_top)
        ws.cell(row=3, column=8).value = '[Add New], [Change Request], [Fix], [Update]'
        apply_style(ws.cell(row=3, column=8), cal, alignment=cal_center_top)

        ws.cell(row=4, column=7).value = 'Actions'
        apply_style(ws.cell(row=4, column=7), cal_bold, alignment=cal_center_top)
        ws.cell(row=4, column=8).value = 'A/M/D - Add/Modify/Delete'
        apply_style(ws.cell(row=4, column=8), cal, alignment=cal_center_top)

        # Row 5 empty
        # Row 6: Header
        headers = ['VERSION', 'CHANGE DESCRIPTION', 'ACTIONS', 'NOTES']
        for i, h in enumerate(headers):
            cell = ws.cell(row=6, column=i+1)
            cell.value = h
            apply_style(cell, cal_bold, alignment=cal_center)

        # Row 7: Interface List section
        ws.cell(row=7, column=1).value = '-'
        ws.cell(row=7, column=2).value = 'Interface List'
        ws.cell(row=7, column=3).value = '-'
        ws.cell(row=7, column=4).value = '-'
        for c in range(1, 5):
            apply_style(ws.cell(row=7, column=c), cal_bold, alignment=cal_center)

        # Rest of existing entries (rows 8-16)
        existing = [
            ('0.1.0', '[Add New]: Initialize the Software Component Interaction ', 'A', '-'),
            ('0.2.0', '[Add New]: Initialize the API Specifications', 'A ', '-'),
            ('1.0.0', '[Add New]: Add Distance Sensor class\n[Update]: reedit the Inter-Class Relationship', 'A/M', '-'),
        ]
        for i, (v, d, a, n) in enumerate(existing):
            r = 8 + i
            ws.cell(row=r, column=1).value = v
            ws.cell(row=r, column=2).value = d
            ws.cell(row=r, column=3).value = a
            ws.cell(row=r, column=4).value = n
            apply_style(ws.cell(row=r, column=1), cal, alignment=cal_center)
            apply_style(ws.cell(row=r, column=2), cal, alignment=cal_left_top)
            apply_style(ws.cell(row=r, column=3), cal, alignment=cal_center)
            apply_style(ws.cell(row=r, column=4), cal_bold if n == '-' else cal, alignment=cal_center)

        # Row 11-16: v1.1.0 with merged cell A11:A16
        # Unmerge and re-merge
        try:
            ws.unmerge_cells('A11:A16')
        except:
            pass
        ws.merge_cells('A11:A16')
        ws.cell(row=11, column=1).value = '1.1.0'
        apply_style(ws.cell(row=11, column=1), cal, alignment=cal_center)

        v110_entries = [
            ('[Update]: Add APIs in SensorDaemon; Update ByteTrack note in inference loop in FRT', 'M', '-'),
            ('[Fix] [1.SoftwareComponentInteraction] [FRTApp, SensorDaemon, DBDaemon] Changed', 'M', 'Resolves CONFLICT-04 (ZeroMQ vs D-Bus)'),
            ('[Fix] [2.APISpecifications] [DBDaemon] Renamed duplicate process_environment_eve', 'M', 'Resolves CONFLICT-02 (DbMain duplicate methods & orphaned row)'),
            ('[Update] [2.APISpecifications] [FRTApp, SensorDaemon] Renamed SdbusInterface to', 'M', 'Resolves CONFLICT-05 (Duplicate SdbusInterface names)'),
            ('[Fix] [2.APISpecifications] [SensorDaemon] Changed Member Type of Sht31Driver::d', 'M', 'Resolves CONFLICT-03 (Typo in MemberType)'),
            ('[Update] [0. Overview] [Document Meta] Updated Document Version to 1.1.0 to refl', 'M', 'Resolves CONFLICT-08 (Version Sync)'),
        ]
        for i, (d, a, n) in enumerate(v110_entries):
            r = 12 + i
            ws.cell(row=r, column=2).value = d
            ws.cell(row=r, column=3).value = a
            ws.cell(row=r, column=4).value = n
            apply_style(ws.cell(row=r, column=1), cal, alignment=cal_center)
            apply_style(ws.cell(row=r, column=2), cal, alignment=cal_left_top)
            apply_style(ws.cell(row=r, column=3), cal, alignment=cal_center)
            apply_style(ws.cell(row=r, column=4), cal, alignment=cal_center)

        # Now add v1.2.0 entry at row 17 (after v1.1.0 entries)
        r = 17
        ws.cell(row=r, column=1).value = '1.2.0'
        apply_style(ws.cell(row=r, column=1), cal, alignment=cal_center)

        v120_desc = (
            '[Update]: [1.SoftwareComponentInteraction] [All] Added RecommendDaemon, RecipeExtractor, C TFLite Reader\n'
            '[Update]: [2.APISpecifications] [All] Added RecommendDaemon, RecipeExtractor API tables\n'
            '[Update]: [3.Inter-Class Relationships] [All] Added IPC links for new components\n'
            '[Fix]: [2.APISpecifications] [FRTApp] Updated DeepSORT -> ByteTrack, ultralytics -> tflite-runtime\n'
            '[Update]: [0.Overview] [Meta] Updated Document Version to 1.2.0'
        )
        ws.cell(row=r, column=2).value = v120_desc
        apply_style(ws.cell(row=r, column=2), cal, alignment=cal_left_top)
        ws.cell(row=r, column=3).value = 'M'
        apply_style(ws.cell(row=r, column=3), cal, alignment=cal_center)
        ws.cell(row=r, column=4).value = 'Align with current source architecture'
        apply_style(ws.cell(row=r, column=4), cal, alignment=cal_center)

    # ----- Fix 2: API table formatting (RecommendDaemon & RecipeExtractor) -----
    ws2 = wb['2.APISpecifications']

    # Find and fix Bảng 5: rows 273-303
    # Find and fix Bảng 6: rows 304-317
    for section_start in [273, 304]:
        # Section header (Bảng X)
        cell = ws2.cell(row=section_start, column=1)
        style_api_section_header(cell)
        # Also style columns 2-7
        for c in range(2, 8):
            style_api_section_header(ws2.cell(row=section_start, column=c))

        # Header row (Package, Class, Member Type...)
        hr = section_start + 1
        for c in range(1, 8):
            style_api_header_row(ws2.cell(row=hr, column=c))

        # Data rows
        dr = section_start + 2
        while dr <= ws2.max_row:
            val = ws2.cell(row=dr, column=1).value
            if val is None or str(val).startswith('Bảng') or str(val).startswith('Package'):
                break
            for c in range(1, 8):
                cell = ws2.cell(row=dr, column=c)
                is_bold = (c == 1 and cell.value is not None)
                style_api_data_row(cell, is_first_col_bold=is_bold)
                # Set alignment
                if c == 1:
                    cell.alignment = gst_center
                elif c in (3, 5, 7):
                    cell.alignment = gst_center
                else:
                    cell.alignment = gst_left
            dr += 1

    # ----- Fix 3: ComponentInteraction new rows -----
    ws3 = wb['1.SoftwareComponentInteraction']
    for row in range(1, ws3.max_row + 1):
        v = str(ws3.cell(row=row, column=1).value or '')
        if 'RecommendDaemon' in v or 'RecipeExtractor' in v:
            for c in range(1, 6):
                cell = ws3.cell(row=row, column=c)
                apply_style(cell, gst, alignment=gst_left, border=thin_border)
                if c == 1:
                    cell.font = gst_bold

    # ----- Fix 4: Inter-Class Relationships new rows -----
    ws4 = wb['3. Inter-Class Relationships']
    for row in range(1, ws4.max_row + 1):
        v = str(ws4.cell(row=row, column=1).value or '')
        if v in ('RecommendDaemon', 'RecipeExtractor', 'UI', 'FRTApp'):
            # Check if this is a new row by looking at column 7
            desc = str(ws4.cell(row=row, column=7).value or '')
            if 'RecommendDaemon' in desc or 'RecipeExtractor' in desc or 'UI' in desc:
                for c in range(1, 12):
                    cell = ws4.cell(row=row, column=c)
                    apply_style(cell, gst, alignment=gst_left, border=thin_border)
                    if c == 1:
                        cell.font = gst_bold

    wb.save(path)
    print(f'Fixed: {path}')

# ============================================================
# Fix SWE v1.3.0
# ============================================================
def fix_swe():
    path = os.path.join(DOCS, 'FSS_SoftwareEngineering_v1.3.0.xlsx')
    wb = load_workbook(path)

    # Fix ChangeLog entry format
    ws = wb['x. ChangeLog']
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == '1.3.0':
            for c in range(1, 5):
                cell = ws.cell(row=row, column=c)
                if c == 2:
                    apply_style(cell, cal, alignment=cal_left_top)
                else:
                    apply_style(cell, cal, alignment=cal_center)
            break

    # Fix new requirement rows formatting (rows 32+)
    ws2 = wb['1. SoftwareReqSpec']
    # Sample original formatting from existing requirement rows
    for row in range(32, ws2.max_row + 1):
        val = ws2.cell(row=row, column=1).value
        if val is None and ws2.cell(row=row, column=3).value == '-':
            # Heading row
            for c in range(1, 14):
                cell = ws2.cell(row=row, column=c)
                if cell.value:
                    apply_style(cell, cal_bold, alignment=cal_left_top)
        elif ws2.cell(row=row, column=3).value:
            # Data row
            for c in range(1, 14):
                cell = ws2.cell(row=row, column=c)
                if cell.value:
                    if c in (1, 3, 4, 9, 10, 11, 12):
                        apply_style(cell, cal, alignment=cal_center)
                    else:
                        apply_style(cell, cal, alignment=cal_left_top)

    wb.save(path)
    print(f'Fixed: {path}')

# ============================================================
# Fix SYS v1.2.0
# ============================================================
def fix_sys():
    path = os.path.join(DOCS, 'FSS_SystemEngineering_v1.2.0.xlsx')
    wb = load_workbook(path)

    # Fix ChangeLog entry format (row 28)
    ws = wb['ChangeLog']
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == '1.2.0':
            for c in range(1, 5):
                cell = ws.cell(row=row, column=c)
                if c == 2:
                    apply_style(cell, cal, alignment=cal_left_top)
                else:
                    apply_style(cell, cal, alignment=cal_center)
            break

    # Fix new Elements (rows 14-16)
    ws2 = wb['2. Elements']
    for row in range(14, ws2.max_row + 1):
        v = ws2.cell(row=row, column=3).value
        if v and 'RecommendDaemon' in str(v):
            for c in range(1, 8):
                cell = ws2.cell(row=row, column=c)
                apply_style(cell, cal, alignment=cal_left_top, border=thin_border)
            break

    # Fix new Interfaces (rows 14-17)
    ws3 = wb['3. Interfaces']
    for row in range(14, ws3.max_row + 1):
        v = ws3.cell(row=row, column=1).value
        if v and 'SYS.IF.INT' in str(v):
            for c in range(1, 7):
                cell = ws3.cell(row=row, column=c)
                apply_style(cell, cal, alignment=cal_left_top, border=thin_border)

    # Fix new ReqSpec entries
    ws4 = wb['1. ReqSpec']
    for row in range(11, ws4.max_row + 1):
        v = ws4.cell(row=row, column=2).value
        if v and 'SYS.FUNC.AI.02' in str(v):
            for c in range(1, 9):
                cell = ws4.cell(row=row, column=c)
                apply_style(cell, cal, alignment=cal_left_top)

    wb.save(path)
    print(f'Fixed: {path}')

# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    fix_sdd()
    fix_swe()
    fix_sys()
    print('\nAll formatting fixes applied.')
