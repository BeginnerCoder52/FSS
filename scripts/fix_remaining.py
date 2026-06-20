#!/usr/bin/env python3
"""Fix remaining formatting issues in the updated xlsx docs."""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

DOCS = os.path.join(os.path.dirname(__file__), '..', 'docs')

cal = Font(name='Calibri', size=11)
cal_bold = Font(name='Calibri', size=11, bold=True)
cal_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
cal_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# ============ SWE v1.3.0 ============
path = os.path.join(DOCS, 'FSS_SoftwareEngineering_v1.3.0.xlsx')
wb = load_workbook(path)
ws = wb['x. ChangeLog']

# Move v1.3.0 entry from row 1 to row 10 (after v0.2.0 at row 9)
if ws.cell(row=1, column=1).value == '1.3.0':
    entry = {c: ws.cell(row=1, column=c).value for c in range(1, 5)}
    # Clear row 1 and rows 2-5 (instructions)
    for r in range(1, 6):
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None

    # Restore rows 1-5 (instructions + header)
    ws.cell(row=1, column=6).value = 'Change Description:'
    ws.cell(row=2, column=7).value = 'Must have [Reason] [Sheet changed][Sections changed][Details]'
    ws.cell(row=3, column=7).value = 'Reason'
    ws.cell(row=3, column=8).value = '[Add New], [Change Request], [Fix], [Update]'
    ws.cell(row=4, column=7).value = 'Actions'
    ws.cell(row=4, column=8).value = 'A/M/D - Add/Modify/Delete'

    for r in range(1, 6):
        for c in range(7, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = cal if r > 1 else cal_bold
            cell.alignment = cal_center

    # Row 6: Header
    headers = ['VERSION', 'CHANGE DESCRIPTION', 'ACTIONS', 'NOTES']
    for i, h in enumerate(headers):
        cell = ws.cell(row=6, column=i+1)
        cell.value = h
        cell.font = cal_bold
        cell.alignment = cal_center

    # Row 7: Section header
    ws.cell(row=7, column=1).value = '-'
    ws.cell(row=7, column=2).value = 'Software Requirement Specifications'
    ws.cell(row=7, column=3).value = '-'
    ws.cell(row=7, column=4).value = '-'
    for c in range(1, 5):
        cell = ws.cell(row=7, column=c)
        cell.font = cal_bold
        cell.alignment = cal_center

    # Row 8: v0.1.0
    ws.cell(row=8, column=1).value = '0.1.0'
    ws.cell(row=8, column=2).value = '[Add New]: Initial the Software Requirement Specifications.'
    ws.cell(row=8, column=3).value = 'A'
    ws.cell(row=8, column=4).value = '-'
    for c in range(1, 5):
        cell = ws.cell(row=8, column=c)
        cell.font = cal
        cell.alignment = cal_center if c != 2 else cal_left_top

    # Row 9: v0.2.0
    ws.cell(row=9, column=1).value = '0.2.0'
    ws.cell(row=9, column=2).value = (
        '[Update]: Updated Software Requirements to align with new hardware (USB Camera) '
        'and new FRT Pipeline v1.2 logic.'
    )
    ws.cell(row=9, column=3).value = 'M'
    ws.cell(row=9, column=4).value = 'Focus on SWS.FUNC.CAM, SWS.FUNC.AI, and SWS.FUNC.DB.'
    for c in range(1, 5):
        cell = ws.cell(row=9, column=c)
        cell.font = cal
        cell.alignment = cal_center if c != 2 else cal_left_top

    # Row 10: v1.3.0 (our new entry)
    ws.cell(row=10, column=1).value = entry[1]
    ws.cell(row=10, column=2).value = (
        '[Add New]: Added RecommendDaemon requirements (SW.REC.01-04)\n'
        '[Add New]: Added RecipeExtractor requirements (SW.EXT.01-03)\n'
        '[Add New]: Added C TFLite Reader requirements (SW.CTL.01-03)\n'
        '[Add New]: Added D-Bus IPC requirements (SW.IPC.01-02)\n'
        '[Update]: Added Test Framework requirement (SW.TST.01)\n'
        '[Update]: Updated Document Version to 1.3.0'
    )
    ws.cell(row=10, column=3).value = 'A'
    ws.cell(row=10, column=4).value = 'Align with SDD v1.2.0 and new architecture'
    for c in range(1, 5):
        cell = ws.cell(row=10, column=c)
        cell.font = cal
        cell.alignment = cal_center if c != 2 else cal_left_top

wb.save(path)
print(f'Fixed SWE ChangeLog')

# ============ SYS v1.2.0 ============
path = os.path.join(DOCS, 'FSS_SystemEngineering_v1.2.0.xlsx')
wb = load_workbook(path)

# Fix ChangeLog entry at row 28 - apply proper font
ws = wb['ChangeLog']
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == '1.2.0':
        # Fix all 1.2.0 entries
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.font = cal
            if c == 2:
                cell.alignment = cal_left_top
            else:
                cell.alignment = cal_center

# Fix Elements rows 17-19
ws2 = wb['2. Elements']
for row in range(17, 20):
    for c in range(1, 8):
        cell = ws2.cell(row=row, column=c)
        cell.font = cal
        cell.alignment = cal_left_top
        cell.border = thin_border

# Also fix row numbers for Elements
for i, no_val in enumerate([11, 12, 13]):
    ws2.cell(row=17 + i, column=1).value = no_val

wb.save(path)
print(f'Fixed SYS ChangeLog and Elements')

print('\nDone.')
